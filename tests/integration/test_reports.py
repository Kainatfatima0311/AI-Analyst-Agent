"""Dashboard figures, saved reports and the files they export to.

Against a real database, because every claim here is about SQL or about a file format. The
dashboard's arithmetic is the kind that looks right until somebody asks a question mid-run and the
success rate drops; the report snapshot's whole purpose is that it does *not* change when the
system behind it does. Neither can be checked against a mock.
"""

from __future__ import annotations

import io
import uuid
import zipfile
from typing import Any

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from analyst_agent.api.main import app
from analyst_agent.db import repository as repo
from analyst_agent.reports.export import to_excel, to_pdf
from analyst_agent.reports.snapshot import build_snapshot, default_name, safe_filename

ANSWER: dict[str, Any] = {
    "conclusion": "Revenue fell 32% in March 2018, driven by a shift away from premium categories.",
    "confidence": "medium",
    "caveats": ["Excludes cancelled orders, per the approved definition."],
    "refuted": ["Delivery delays: lateness was flat across seller states."],
}


@pytest.fixture
def client(rw_dsn: str):
    with TestClient(app) as test_client:
        yield test_client


@pytest.fixture
def answered_run(rw_dsn: str) -> uuid.UUID:
    """A finished run with evidence, a material finding and two tested explanations."""
    run_id = repo.create_run("Why did revenue drop in March 2018?", requested_by="tests")

    good = repo.record_sql_audit(
        run_id=run_id,
        purpose="monthly net revenue [revenue@v1]",
        sql_text="SELECT 1",
        verdict="allowed",
        reasons=[],
        rewritten_sql="SELECT 1 LIMIT 5000",
        referenced_objects=["analytics.orders"],
        sensitive_columns=[],
        estimated_cost=120.0,
    )
    repo.mark_sql_executed(good, row_count=12, truncated=False, duration_ms=31)

    second = repo.record_sql_audit(
        run_id=run_id,
        purpose="premium category share",
        sql_text="SELECT 2",
        verdict="allowed",
        reasons=[],
        rewritten_sql=None,
        referenced_objects=["analytics.order_items"],
        sensitive_columns=[],
        estimated_cost=90.0,
    )
    repo.mark_sql_executed(second, row_count=6, truncated=False, duration_ms=18)

    # A refused statement, so the report has to prove it carries those too.
    repo.record_sql_audit(
        run_id=run_id,
        purpose="attempted deletion",
        sql_text="DROP TABLE analytics.orders",
        verdict="rejected",
        reasons=["not_a_select: the statement is a Drop, not a SELECT"],
        rewritten_sql=None,
        referenced_objects=[],
        sensitive_columns=[],
        estimated_cost=None,
    )

    finding_id = repo.record_finding(
        run_id=run_id,
        statement="Net revenue fell 32% in 2018-03.",
        material=True,
        evidence_query_ids=[good],
    )
    supported = repo.record_hypothesis(
        run_id=run_id,
        finding_id=finding_id,
        statement="The premium category share collapsed.",
        test_design="compare premium share of revenue, 2018-02 against 2018-03",
    )
    repo.update_hypothesis(
        supported, status="supported", reasoning="share fell from 0.28 to 0.11",
        test_query_ids=[second],
    )
    refuted = repo.record_hypothesis(
        run_id=run_id,
        finding_id=finding_id,
        statement="Delivery delays pushed cancellations up.",
        test_design="late-delivery rate by seller state, month over month",
    )
    repo.update_hypothesis(
        refuted, status="refuted", reasoning="lateness was flat", test_query_ids=[second]
    )

    repo.finish_run(
        run_id,
        "completed",
        answer={**ANSWER, "evidence": [{"query_id": str(good)}, {"query_id": str(second)}]},
    )
    return run_id


# --- the dashboard -------------------------------------------------------------


def test_the_dashboard_counts_analyses_and_reports(client: TestClient, answered_run) -> None:
    """Counted as deltas: the database is shared across tests, so absolutes would be brittle."""
    before = client.get("/v1/dashboard/summary").json()["totals"]
    assert before["analyses"] >= 1
    assert before["completed"] >= 1

    client.post("/v1/reports", json={"run_id": str(answered_run), "name": "March revenue"})
    after = client.get("/v1/dashboard/summary").json()["totals"]
    assert after["saved_reports"] == before["saved_reports"] + 1
    assert after["analyses"] == before["analyses"], "saving a report is not a new analysis"


def test_the_success_rate_is_over_finished_runs_only(client: TestClient, answered_run) -> None:
    """Counting an in-flight run as a failure would drop the rate every time somebody asks."""
    before = client.get("/v1/dashboard/summary").json()["outcomes"]
    repo.create_run("a question still in flight")
    after = client.get("/v1/dashboard/summary").json()

    assert after["outcomes"]["success_rate"] == before["success_rate"]
    assert after["totals"]["in_flight"] >= 1
    assert after["totals"]["analyses"] > before["finished"]


def test_a_failed_run_moves_the_rates(client: TestClient, answered_run) -> None:
    failed = repo.create_run("a question that fails")
    repo.finish_run(failed, "failed", error={"type": "Boom", "message": "no"})
    outcomes = client.get("/v1/dashboard/summary").json()["outcomes"]
    assert outcomes["failure_rate"] and outcomes["failure_rate"] > 0
    assert round(outcomes["success_rate"] + outcomes["failure_rate"], 1) <= 100


def test_the_dashboard_lists_recent_questions_newest_first(client: TestClient, answered_run) -> None:
    repo.create_run("the newest question")
    questions = client.get("/v1/dashboard/summary").json()["recent_questions"]
    assert questions[0]["question"] == "the newest question"


def test_recent_insights_come_from_findings(client: TestClient, answered_run) -> None:
    """The closest thing the system has to "what it noticed"."""
    insights = client.get("/v1/dashboard/summary").json()["recent_insights"]
    assert any("Net revenue fell 32%" in item["statement"] for item in insights)
    assert all(item["question"] for item in insights), "an insight without its question is orphaned"


def test_most_used_metrics_come_from_the_tool_calls(client: TestClient, answered_run) -> None:
    """Read from where a metric is named, not parsed out of SQL text.

    `metric_query` is the only place an approved definition is invoked *by name*, so the count is
    a fact rather than a guess at what a statement was computing.
    """
    for _ in range(3):
        repo.record_tool_call(
            run_id=answered_run,
            tool="metric_query",
            arguments={"metric": "revenue", "purpose": "monthly"},
            ok=True,
            step_id=None,
        )
    repo.record_tool_call(
        run_id=answered_run, tool="metric_query", arguments={"metric": "aov"}, ok=True,
        step_id=None,
    )
    repo.record_tool_call(
        run_id=answered_run, tool="sql_runner", arguments={"sql": "SELECT 1"}, ok=True,
        step_id=None,
    )

    metrics = {m["metric"]: m["uses"] for m in client.get("/v1/dashboard/summary").json()["top_metrics"]}
    assert metrics["revenue"] >= 3
    assert metrics["aov"] >= 1
    assert "sql_runner" not in metrics, "only metric_query names a definition"


def test_the_dashboard_bounds_its_lists(client: TestClient, answered_run) -> None:
    body = client.get("/v1/dashboard/summary?recent=1").json()
    assert len(body["recent_questions"]) == 1


def test_the_dashboard_survives_an_empty_database(client: TestClient, rw_dsn: str) -> None:
    """An empty state has to render, not divide by zero."""
    body = client.get("/v1/dashboard/summary").json()
    assert body["totals"]["analyses"] >= 0
    assert body["outcomes"]["success_rate"] is None or body["outcomes"]["success_rate"] >= 0


# --- saving, renaming, deleting -----------------------------------------------


def test_a_report_can_be_saved_read_renamed_and_deleted(client: TestClient, answered_run) -> None:
    created = client.post(
        "/v1/reports",
        json={"run_id": str(answered_run), "name": "March revenue", "saved_by": "tests"},
    )
    assert created.status_code == 201
    report_id = created.json()["report_id"]

    listed = client.get("/v1/reports").json()
    mine = next(r for r in listed if r["report_id"] == report_id)
    assert listed[0]["report_id"] == report_id, "newest first"
    assert mine["name"] == "March revenue"
    assert mine["queries"] == 2, "the list counts without loading the snapshot"

    renamed = client.patch(f"/v1/reports/{report_id}", json={"name": "Q1 revenue review"})
    assert renamed.status_code == 200
    assert renamed.json()["name"] == "Q1 revenue review"
    assert client.get(f"/v1/reports/{report_id}").json()["name"] == "Q1 revenue review"

    assert client.delete(f"/v1/reports/{report_id}").status_code == 204
    assert client.get(f"/v1/reports/{report_id}").status_code == 404
    assert report_id not in [r["report_id"] for r in client.get("/v1/reports").json()]


def test_a_report_takes_the_question_as_its_name_when_none_is_given(
    client: TestClient, answered_run
) -> None:
    created = client.post("/v1/reports", json={"run_id": str(answered_run)})
    assert created.json()["name"] == "Why did revenue drop in March 2018?"


def test_a_run_with_no_answer_cannot_be_saved(client: TestClient, rw_dsn: str) -> None:
    """A report whose body is empty is a filename, and it would sit in the list looking real."""
    unfinished = repo.create_run("still going")
    refused = client.post("/v1/reports", json={"run_id": str(unfinished)})
    assert refused.status_code == 409
    assert "nothing to save" in refused.json()["detail"]


def test_saving_an_unknown_run_is_a_404(client: TestClient, rw_dsn: str) -> None:
    assert client.post("/v1/reports", json={"run_id": str(uuid.uuid4())}).status_code == 404


def test_renaming_and_deleting_an_unknown_report_are_404s(client: TestClient, rw_dsn: str) -> None:
    missing = uuid.uuid4()
    assert client.patch(f"/v1/reports/{missing}", json={"name": "x"}).status_code == 404
    assert client.delete(f"/v1/reports/{missing}").status_code == 404


def test_a_blank_name_is_refused(client: TestClient, answered_run) -> None:
    """The database refuses it too — a report nobody can find again is not saved."""
    report_id = client.post("/v1/reports", json={"run_id": str(answered_run)}).json()["report_id"]
    assert client.patch(f"/v1/reports/{report_id}", json={"name": "   "}).status_code == 422


def test_deleting_the_run_deletes_its_reports(client: TestClient, answered_run) -> None:
    """A report pointing at a run that no longer exists would be a dangling artefact."""
    report_id = client.post("/v1/reports", json={"run_id": str(answered_run)}).json()["report_id"]
    with repo.rw_conn() as conn, conn.cursor() as cur:
        cur.execute("DELETE FROM agent.runs WHERE run_id = %s", (answered_run,))
    assert client.get(f"/v1/reports/{report_id}").status_code == 404


# --- the snapshot --------------------------------------------------------------


def test_the_snapshot_carries_everything_a_reader_needs(client: TestClient, answered_run) -> None:
    """Question, answer, charts, SQL evidence, definition versions, timestamp."""
    snapshot = client.post("/v1/reports", json={"run_id": str(answered_run)}).json()["snapshot"]

    assert snapshot["question"] == "Why did revenue drop in March 2018?"
    assert "Revenue fell 32%" in snapshot["answer"]["conclusion"]
    assert snapshot["answer"]["refuted"], "what was ruled out travels with the answer"
    assert snapshot["confidence"]["score"] >= 0
    assert snapshot["confidence"]["factors"], "a score without its factors is decoration"
    assert snapshot["saved_at"], "a report has to say when it was taken"
    assert [e["sql"] for e in snapshot["evidence"]] == ["SELECT 1 LIMIT 5000", "SELECT 2"]
    assert snapshot["evidence"][0]["definition_version"] == "revenue@v1"
    assert any(m["metric"] == "revenue" for m in snapshot["metrics_used"])
    assert len(snapshot["findings"][0]["hypotheses"]) == 2


def test_the_snapshot_keeps_the_queries_that_never_ran(client: TestClient, answered_run) -> None:
    """A report showing only what ran would hide the half a reviewer asks about."""
    snapshot = client.post("/v1/reports", json={"run_id": str(answered_run)}).json()["snapshot"]
    refused = [q for q in snapshot["queries_considered"] if q["verdict"] == "rejected"]
    assert refused, "the blocked attempt is part of the record"
    assert "DROP TABLE" in refused[0]["sql"]
    assert refused[0]["reasons"]


def test_a_saved_report_does_not_change_when_the_run_does(client: TestClient, answered_run) -> None:
    """The whole reason a report is a snapshot rather than a pointer.

    Somebody re-opening a report from March must not find different figures under the same name
    because a definition was revised in between.
    """
    report_id = client.post("/v1/reports", json={"run_id": str(answered_run)}).json()["report_id"]
    before = client.get(f"/v1/reports/{report_id}").json()["snapshot"]

    repo.finish_run(answered_run, "completed", answer={**ANSWER, "conclusion": "Something else."})
    after = client.get(f"/v1/reports/{report_id}").json()["snapshot"]

    assert after["answer"]["conclusion"] == before["answer"]["conclusion"]
    assert "Something else." not in after["answer"]["conclusion"]
    # The live run *has* moved on, and the report still points at it.
    assert client.get(f"/v1/runs/{answered_run}").json()["answer"]["conclusion"] == "Something else."


def test_a_dangling_citation_is_dropped_rather_than_rendered_empty(rw_dsn: str) -> None:
    trace = {"run": {"question": "q"}, "queries": [], "findings": [], "hypotheses": []}
    snapshot = build_snapshot(trace, {"conclusion": "c", "evidence": [{"query_id": "missing"}]})
    assert snapshot["evidence"] == []


# --- exports -------------------------------------------------------------------


def test_the_pdf_export_is_a_pdf_and_carries_the_evidence(client: TestClient, answered_run) -> None:
    report_id = client.post(
        "/v1/reports", json={"run_id": str(answered_run), "name": "March revenue"}
    ).json()["report_id"]

    response = client.get(f"/v1/reports/{report_id}/export.pdf")
    assert response.status_code == 200
    assert response.headers["content-type"] == "application/pdf"
    assert "March-revenue.pdf" in response.headers["content-disposition"]
    assert response.content.startswith(b"%PDF-"), "a real PDF, not an error page"
    assert int(response.headers["content-length"]) == len(response.content)
    # A one-page PDF would mean the evidence section never rendered.
    assert response.content.count(b"/Type /Page") >= 2 or b"/Count 2" in response.content


def test_the_pdf_survives_markup_in_the_answer(rw_dsn: str) -> None:
    """reportlab reads its own mini-markup, so `<` in a conclusion must not break the build."""
    report = {
        "name": "angle brackets",
        "snapshot": {
            "question": "does 3 < 4?",
            "answer": {"conclusion": "yes: 3 < 4 & 5 > 2", "caveats": [], "refuted": []},
            "confidence": {"score": 50, "band": "medium", "factors": []},
            "evidence": [],
            "queries_considered": [],
            "findings": [],
            "charts": [],
            "metrics_used": [],
        },
    }
    assert to_pdf(report).startswith(b"%PDF-")


def test_the_excel_export_has_a_sheet_per_kind_of_thing(client: TestClient, answered_run) -> None:
    report_id = client.post("/v1/reports", json={"run_id": str(answered_run)}).json()["report_id"]
    response = client.get(f"/v1/reports/{report_id}/export.xlsx")
    assert response.status_code == 200
    assert "spreadsheetml" in response.headers["content-type"]

    book = load_workbook(io.BytesIO(response.content))
    assert book.sheetnames == [
        "Summary", "Confidence", "Findings", "Evidence", "All queries", "Definitions", "Charts",
    ]

    evidence = book["Evidence"]
    assert [cell.value for cell in evidence[1]][:2] == ["Purpose", "Definition"]
    sql_column = [row[5].value for row in evidence.iter_rows(min_row=2)]
    assert "SELECT 1 LIMIT 5000" in sql_column, "the SQL is a column, so it survives copy-paste"
    assert evidence.freeze_panes == "A2"


def test_the_excel_export_keeps_the_confidence_factors(client: TestClient, answered_run) -> None:
    report_id = client.post("/v1/reports", json={"run_id": str(answered_run)}).json()["report_id"]
    book = load_workbook(io.BytesIO(client.get(f"/v1/reports/{report_id}/export.xlsx").content))
    labels = [row[0].value for row in book["Confidence"].iter_rows(min_row=2)]
    assert any("quer" in str(label) for label in labels)


def test_an_empty_sheet_still_has_its_headers(rw_dsn: str) -> None:
    """A reader who opens "Findings" and sees nothing has learnt there were none."""
    book = load_workbook(
        io.BytesIO(
            to_excel({"name": "bare", "snapshot": {"question": "q", "answer": {}}})
        )
    )
    assert next(cell.value for cell in book["Findings"][1]) == "Finding"


def test_the_excel_file_is_a_real_workbook(client: TestClient, answered_run) -> None:
    """xlsx is a zip; a truncated write would still return 200 with unusable bytes."""
    report_id = client.post("/v1/reports", json={"run_id": str(answered_run)}).json()["report_id"]
    payload = client.get(f"/v1/reports/{report_id}/export.xlsx").content
    with zipfile.ZipFile(io.BytesIO(payload)) as archive:
        assert archive.testzip() is None
        assert "xl/workbook.xml" in archive.namelist()


def test_exporting_an_unknown_report_is_a_404(client: TestClient, rw_dsn: str) -> None:
    missing = uuid.uuid4()
    assert client.get(f"/v1/reports/{missing}/export.pdf").status_code == 404
    assert client.get(f"/v1/reports/{missing}/export.xlsx").status_code == 404


def test_a_chart_with_no_stored_image_is_a_404_not_an_empty_file(
    client: TestClient, answered_run
) -> None:
    """An empty PNG download looks like a broken image; a 404 says what happened."""
    query_id = repo.get_trace(answered_run)["queries"][0]["query_id"]
    chart_id = repo.record_chart(
        run_id=answered_run,
        query_id=query_id,
        chart_type="line",
        spec={"data": [], "layout": {}},
        title="no image",
        png=None,
    )
    response = client.get(f"/v1/charts/{chart_id}/export.png")
    assert response.status_code == 404
    assert "no stored image" in response.json()["detail"]


def test_a_stored_chart_downloads_as_a_png(client: TestClient, answered_run) -> None:
    png = b"\x89PNG\r\n\x1a\n" + b"0" * 64
    query_id = repo.get_trace(answered_run)["queries"][0]["query_id"]
    chart_id = repo.record_chart(
        run_id=answered_run,
        query_id=query_id,
        chart_type="line",
        spec={"data": [], "layout": {}},
        title="Revenue by month",
        png=png,
    )
    response = client.get(f"/v1/charts/{chart_id}/export.png")
    assert response.status_code == 200
    assert response.headers["content-type"] == "image/png"
    assert "Revenue-by-month.png" in response.headers["content-disposition"]
    assert response.content == png, "the image stored when the chart was built, byte for byte"


# --- naming --------------------------------------------------------------------


def test_a_long_question_is_trimmed_at_a_word() -> None:
    """"Why did revenue drop in Marc" reads as a bug."""
    name = default_name("Why did revenue drop in March 2018 across every product category " * 2)
    assert len(name) <= 71
    assert not name.rstrip("…").endswith(" ")
    assert "…" in name


def test_a_filename_is_safe_for_any_filesystem() -> None:
    assert safe_filename('Q1: revenue/margin "review"', "pdf") == "Q1-revenue-margin-review.pdf"
    assert safe_filename("", "xlsx") == "report.xlsx"
    assert safe_filename("...", "pdf") == "report.pdf"


# --- the answer carries the score ---------------------------------------------


def test_a_finished_run_reports_a_confidence_score(client: TestClient, answered_run) -> None:
    """The spec's headline: every final answer includes a score, not only a band."""
    answer = client.get(f"/v1/runs/{answered_run}").json()["answer"]
    assert 0 <= answer["confidence_score"] <= 100
    assert answer["confidence"] == "medium", "the agent's own band is still reported"
    factors = answer["confidence_detail"]["factors"]
    evidence = next(f for f in factors if f["key"] == "evidence")
    assert evidence["label"] == "2 queries executed"
    assert any(f["key"] == "alternatives" and f["passed"] for f in factors)
