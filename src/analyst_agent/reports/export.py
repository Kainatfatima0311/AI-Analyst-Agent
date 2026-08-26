"""Reports as files: PDF to read, Excel to work with, PNG for a chart on its own.

Both exporters are built from the **snapshot**, never from the live database. That is what makes
an exported file mean the same thing as the report it came from — including six months later,
when the metric definition behind it may have moved on.

Neither format is allowed to drop the parts that make an answer checkable. The rule both follow:
a conclusion, its confidence with the factors behind it, the findings with what was refuted, and
the SQL for every cited number. An export that carried only the conclusion would be the one
artefact in this system that asks to be believed.
"""

from __future__ import annotations

import io
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_LEFT
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    HRFlowable,
    KeepTogether,
    PageBreak,
    Paragraph,
    Preformatted,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

from analyst_agent.observability.logging import get_logger

log = get_logger(__name__)

INK = colors.HexColor("#14162b")
MUTED = colors.HexColor("#5b5f7a")
FAINT = colors.HexColor("#868aa3")
RULE = colors.HexColor("#d3d6e6")
ACCENT = colors.HexColor("#6c5ce7")
GOOD = colors.HexColor("#1f7a4d")
WARN = colors.HexColor("#9a6b00")
BAD = colors.HexColor("#b23c3c")
WASH = colors.HexColor("#f5f6fb")

BAND_COLOUR = {"high": GOOD, "medium": WARN, "low": BAD}


def _styles() -> dict[str, ParagraphStyle]:
    base = getSampleStyleSheet()
    return {
        "title": ParagraphStyle(
            "title", parent=base["Title"], fontSize=17, leading=21, textColor=INK,
            alignment=TA_LEFT, spaceAfter=2,
        ),
        "meta": ParagraphStyle("meta", parent=base["Normal"], fontSize=8.5, textColor=FAINT),
        "h2": ParagraphStyle(
            "h2", parent=base["Heading2"], fontSize=11, leading=14, textColor=INK,
            spaceBefore=13, spaceAfter=5,
        ),
        "body": ParagraphStyle(
            "body", parent=base["Normal"], fontSize=9.5, leading=14.5, textColor=INK,
        ),
        "muted": ParagraphStyle(
            "muted", parent=base["Normal"], fontSize=8.5, leading=12.5, textColor=MUTED,
        ),
        "mono": ParagraphStyle(
            "mono", parent=base["Code"], fontSize=7.4, leading=9.6, textColor=MUTED,
        ),
    }


def to_pdf(report: dict[str, Any]) -> bytes:
    """A report as a PDF somebody could hand to a colleague.

    Ordered the way a reader uses it: the conclusion first, then how confident it is *and why*,
    then what was ruled out, then the evidence. The SQL comes last and complete — it is the part
    nobody reads until they doubt the answer, and the part that has to be there when they do.
    """
    snapshot = report.get("snapshot") or {}
    style = _styles()
    buffer = io.BytesIO()
    document = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=18 * mm,
        rightMargin=18 * mm,
        topMargin=16 * mm,
        bottomMargin=16 * mm,
        title=str(report.get("name") or snapshot.get("question") or "Analysis"),
        author="AI Data Analyst Agent",
    )

    flow: list[Any] = []
    flow.append(Paragraph(_x(report.get("name") or snapshot.get("question")), style["title"]))
    flow.append(
        Paragraph(
            " · ".join(
                bit
                for bit in (
                    _x(snapshot.get("question")),
                    f"asked {_short(snapshot.get('asked_at'))}" if snapshot.get("asked_at") else "",
                    f"saved {_short(snapshot.get('saved_at'))}" if snapshot.get("saved_at") else "",
                    f"run {snapshot.get('run_id', '')[:8]}" if snapshot.get("run_id") else "",
                )
                if bit
            ),
            style["meta"],
        )
    )
    flow.append(Spacer(1, 7))
    flow.append(HRFlowable(width="100%", color=RULE, thickness=0.7, spaceAfter=9))

    answer = snapshot.get("answer") or {}
    flow.append(Paragraph("Conclusion", style["h2"]))
    conclusion = _x(answer.get("conclusion")) or "<i>This run produced no answer.</i>"
    flow.append(Paragraph(conclusion.replace("\n", "<br/>"), style["body"]))

    flow += _confidence_block(snapshot.get("confidence") or {}, style)

    if answer.get("refuted"):
        flow.append(Paragraph("Ruled out", style["h2"]))
        for item in answer["refuted"]:
            flow.append(Paragraph(f"• {_x(item)}", style["body"]))

    if answer.get("caveats"):
        flow.append(Paragraph("Caveats", style["h2"]))
        for item in answer["caveats"]:
            flow.append(Paragraph(f"• {_x(item)}", style["body"]))

    flow += _findings_block(snapshot.get("findings") or [], style)
    flow += _metrics_block(snapshot.get("metrics_used") or [], style)
    flow += _charts_block(snapshot.get("charts") or [], style)

    evidence = snapshot.get("evidence") or []
    considered = snapshot.get("queries_considered") or []
    if evidence or considered:
        flow.append(PageBreak())
        flow.append(Paragraph("Evidence", style["h2"]))
        flow.append(
            Paragraph(
                "Every number in the conclusion traces to one of these queries. Statements the "
                "guard refused are listed too — what the agent tried is usually what a reviewer "
                "wants to know.",
                style["muted"],
            )
        )
        for item in evidence:
            flow.append(Spacer(1, 6))
            head = _x(item.get("purpose"))
            if item.get("definition_version"):
                head += f"  [{_x(item['definition_version'])}]"
            flow.append(
                KeepTogether(
                    [
                        Paragraph(f"<b>{head}</b>", style["body"]),
                        Paragraph(
                            f"{item.get('query_id', '')} · {item.get('row_count')} rows"
                            + (" · truncated" if item.get("truncated") else ""),
                            style["meta"],
                        ),
                        Preformatted(_wrap_sql(item.get("sql", "")), style["mono"]),
                    ]
                )
            )

        refused = [q for q in considered if q.get("verdict") != "allowed" or not q.get("executed")]
        if refused:
            flow.append(Paragraph("Queries that did not run", style["h2"]))
            for item in refused:
                reasons = "; ".join(item.get("reasons") or []) or item.get("verdict", "")
                flow.append(
                    KeepTogether(
                        [
                            Paragraph(f"<b>{_x(item.get('purpose'))}</b>", style["body"]),
                            Paragraph(f"{_x(item.get('verdict'))} — {_x(reasons)}", style["meta"]),
                            Preformatted(_wrap_sql(item.get("sql", "")), style["mono"]),
                        ]
                    )
                )

    document.build(buffer and flow, onFirstPage=_footer, onLaterPages=_footer)
    pdf = buffer.getvalue()
    log.info("report exported", format="pdf", bytes=len(pdf))
    return pdf


def _confidence_block(confidence: dict[str, Any], style: dict[str, ParagraphStyle]) -> list[Any]:
    """The score, and the factors that produced it.

    The factors are the point. A percentage on its own is a claim; a percentage with the four
    things it was computed from is something a reader can disagree with.
    """
    if not confidence:
        return []
    score = confidence.get("score")
    band = str(confidence.get("band") or "")
    rows = [
        [
            Paragraph(f"<b>Confidence: {score}%</b>", style["body"]),
            Paragraph(band.upper(), style["meta"]),
        ]
    ]
    for factor in confidence.get("factors") or []:
        if not factor.get("weight"):
            continue  # not applicable to this run; showing it as a failure would be wrong
        mark = "✓" if factor.get("passed") else "•"
        rows.append(
            [
                Paragraph(f"{mark} {_x(factor.get('label'))}", style["muted"]),
                Paragraph(
                    f"{factor.get('earned', 0):.0f}/{factor.get('weight', 0):.0f}", style["meta"]
                ),
            ]
        )
    if confidence.get("capped_by"):
        rows.append(
            [
                Paragraph(
                    f"Held at the agent's own stated confidence ({_x(confidence['capped_by'])}).",
                    style["muted"],
                ),
                Paragraph("", style["meta"]),
            ]
        )

    table = Table(rows, colWidths=[125 * mm, 24 * mm], hAlign="LEFT")
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), WASH),
                ("BOX", (0, 0), (-1, -1), 0.6, RULE),
                ("LINEBELOW", (0, 0), (-1, 0), 0.6, RULE),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 8),
                ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ("TEXTCOLOR", (1, 0), (1, 0), BAND_COLOUR.get(band, MUTED)),
            ]
        )
    )
    return [Spacer(1, 9), table]


def _findings_block(findings: list[dict[str, Any]], style: dict[str, ParagraphStyle]) -> list[Any]:
    if not findings:
        return []
    flow: list[Any] = [Paragraph("Findings and the explanations tested", style["h2"])]
    for finding in findings:
        label = " [material]" if finding.get("material") else ""
        flow.append(Paragraph(f"<b>{_x(finding.get('statement'))}</b>{label}", style["body"]))
        for hypothesis in finding.get("hypotheses") or []:
            status = str(hypothesis.get("status") or "")
            flow.append(
                Paragraph(
                    f"— {_x(hypothesis.get('statement'))} <b>({_x(status)})</b>"
                    + (f"<br/>{_x(hypothesis.get('reasoning'))}" if hypothesis.get("reasoning") else ""),
                    style["muted"],
                )
            )
        flow.append(Spacer(1, 4))
    return flow


def _metrics_block(metrics: list[dict[str, Any]], style: dict[str, ParagraphStyle]) -> list[Any]:
    if not metrics:
        return []
    flow: list[Any] = [Paragraph("Approved definitions used", style["h2"])]
    flow.append(
        Paragraph(
            "A figure without its definition version is a number whose meaning cannot be "
            "recovered later.",
            style["muted"],
        )
    )
    for metric in metrics:
        version = metric.get("version") or "—"
        flow.append(
            Paragraph(f"• <b>{_x(metric.get('metric'))}</b> · {_x(version)}", style["body"])
        )
    return flow


def _charts_block(charts: list[dict[str, Any]], style: dict[str, ParagraphStyle]) -> list[Any]:
    """Charts are named, not drawn.

    Embedding the PNG would need the image bytes, which live on the chart row rather than in the
    snapshot — and a report that silently rendered a *regenerated* chart would be showing a
    picture that no longer matches the figures beside it. The chart is exported on its own
    endpoint instead, and named here with the query it came from.
    """
    if not charts:
        return []
    flow: list[Any] = [Paragraph("Charts", style["h2"])]
    for chart in charts:
        flow.append(
            Paragraph(
                f"• <b>{_x(chart.get('title'))}</b> ({_x(chart.get('chart_type'))})"
                f" — from query {chart.get('query_id', '')}",
                style["body"],
            )
        )
    flow.append(
        Paragraph("Each chart can be downloaded as a PNG from its own endpoint.", style["muted"])
    )
    return flow


def _footer(canvas: Any, document: Any) -> None:
    canvas.saveState()
    canvas.setFont("Helvetica", 7.5)
    canvas.setFillColor(FAINT)
    canvas.drawString(18 * mm, 10 * mm, "AI Data Analyst Agent — every number traces to a query")
    canvas.drawRightString(A4[0] - 18 * mm, 10 * mm, f"page {document.page}")
    canvas.restoreState()


def to_excel(report: dict[str, Any]) -> bytes:
    """A report as a workbook, one sheet per kind of thing.

    Six sheets rather than one: a reader opening this wants either the numbers or the audit
    trail, and interleaving them produces a sheet that is bad at both. The SQL is a column, not
    a comment, so it survives copy-paste into a query tool.
    """
    snapshot = report.get("snapshot") or {}
    book = Workbook()

    summary = book.active
    summary.title = "Summary"
    answer = snapshot.get("answer") or {}
    confidence = snapshot.get("confidence") or {}
    _rows(
        summary,
        [
            ("Report", report.get("name")),
            ("Question", snapshot.get("question")),
            ("Status", snapshot.get("status")),
            ("Asked", _short(snapshot.get("asked_at"))),
            ("Saved", _short(snapshot.get("saved_at"))),
            ("Run id", snapshot.get("run_id")),
            ("Confidence", f"{confidence.get('score')}% ({confidence.get('band')})"),
            ("Stated confidence", answer.get("stated_confidence")),
            ("Conclusion", answer.get("conclusion")),
        ],
    )
    summary.column_dimensions["A"].width = 20
    summary.column_dimensions["B"].width = 110
    for row in summary.iter_rows(min_col=2, max_col=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    _sheet(
        book,
        "Confidence",
        ["Factor", "Passed", "Earned", "Weight", "Why it matters"],
        [
            [
                factor.get("label"),
                "yes" if factor.get("passed") else "no",
                factor.get("earned"),
                factor.get("weight"),
                factor.get("detail"),
            ]
            for factor in confidence.get("factors") or []
            if factor.get("weight")
        ],
        widths=[52, 9, 9, 9, 70],
    )

    _sheet(
        book,
        "Findings",
        ["Finding", "Material", "Explanation", "Verdict", "Reasoning"],
        [
            [
                finding.get("statement"),
                "yes" if finding.get("material") else "no",
                hypothesis.get("statement"),
                hypothesis.get("status"),
                hypothesis.get("reasoning"),
            ]
            for finding in snapshot.get("findings") or []
            for hypothesis in (finding.get("hypotheses") or [{}])
        ],
        widths=[54, 9, 54, 13, 60],
    )

    _sheet(
        book,
        "Evidence",
        ["Purpose", "Definition", "Rows", "Truncated", "Query id", "SQL"],
        [
            [
                item.get("purpose"),
                item.get("definition_version"),
                item.get("row_count"),
                "yes" if item.get("truncated") else "no",
                item.get("query_id"),
                item.get("sql"),
            ]
            for item in snapshot.get("evidence") or []
        ],
        widths=[46, 16, 8, 11, 38, 100],
    )

    _sheet(
        book,
        "All queries",
        ["Purpose", "Verdict", "Executed", "Rows", "Reasons", "SQL"],
        [
            [
                item.get("purpose"),
                item.get("verdict"),
                "yes" if item.get("executed") else "no",
                item.get("row_count"),
                "; ".join(item.get("reasons") or []),
                item.get("sql"),
            ]
            for item in snapshot.get("queries_considered") or []
        ],
        widths=[46, 12, 11, 8, 46, 100],
    )

    _sheet(
        book,
        "Definitions",
        ["Metric", "Version", "Times computed"],
        [
            [metric.get("metric"), metric.get("version"), metric.get("uses")]
            for metric in snapshot.get("metrics_used") or []
        ],
        widths=[26, 18, 16],
    )

    _sheet(
        book,
        "Charts",
        ["Title", "Type", "Query id", "Chart id"],
        [
            [
                chart.get("title"),
                chart.get("chart_type"),
                chart.get("query_id"),
                chart.get("chart_id"),
            ]
            for chart in snapshot.get("charts") or []
        ],
        widths=[46, 14, 38, 38],
    )

    buffer = io.BytesIO()
    book.save(buffer)
    data = buffer.getvalue()
    log.info("report exported", format="xlsx", bytes=len(data))
    return data


def _rows(sheet: Any, pairs: list[tuple[str, Any]]) -> None:
    for label, value in pairs:
        sheet.append([label, "" if value is None else str(value)])
    for cell in sheet["A"]:
        cell.font = Font(bold=True, color="14162B")


def _sheet(
    book: Workbook, title: str, headers: list[str], rows: list[list[Any]], widths: list[int]
) -> None:
    """One sheet, with a header row that stays visible.

    An empty sheet still gets its headers: a reader who opens "Findings" and sees nothing has
    learnt that there were none, which is different from the sheet being missing.
    """
    sheet = book.create_sheet(title)
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="6C5CE7")
        cell.alignment = Alignment(vertical="center")
    for row in rows:
        sheet.append(["" if value is None else value for value in row])
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.freeze_panes = "A2"
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")


def _wrap_sql(sql: str, width: int = 104) -> str:
    """Hard-wrap SQL so a long line does not run off the page.

    Wrapped rather than shrunk: a statement in 5pt type is present but unreadable, which is the
    worst of both.
    """
    out: list[str] = []
    for line in (sql or "").splitlines() or [""]:
        while len(line) > width:
            cut = line.rfind(" ", 0, width)
            cut = cut if cut > width // 2 else width
            out.append(line[:cut])
            line = "    " + line[cut:].lstrip()
        out.append(line)
    return "\n".join(out) or "(no statement recorded)"


def _x(value: Any) -> str:
    """Escape for reportlab's mini-markup, which treats < and & as markup."""
    return (
        str(value if value is not None else "")
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
    )


def _short(iso: Any) -> str:
    text = str(iso or "")
    return text[:16].replace("T", " ") if text else ""
